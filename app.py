import os
import sqlite3
import csv
import uuid
import re
from flask import Flask, request, jsonify, send_file
from werkzeug.utils import secure_filename
import openpyxl

app = Flask(__name__, static_folder='static', static_url_path='')

# Configuration
DB_PATH = os.path.join(os.path.dirname(__file__), 'data.db')
UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'temp')
ALLOWED_EXTENSIONS = {'csv', 'xlsx', 'xls'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max upload

# Ensure directories exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def get_db_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    # Enable foreign keys
    conn.execute("PRAGMA foreign_keys = ON")
    return conn

def init_db():
    # Database is automatically created on connection.
    # Create the _sys_logical_relations table for virtual relationships
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS _sys_logical_relations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            parent_table TEXT NOT NULL,
            parent_column TEXT NOT NULL,
            child_table TEXT NOT NULL,
            child_column TEXT NOT NULL,
            on_update TEXT DEFAULT 'CASCADE',
            on_delete TEXT DEFAULT 'CASCADE',
            UNIQUE(parent_table, parent_column, child_table, child_column)
        );
    """)
    conn.commit()
    conn.close()

init_db()

# Helper function to check allowed file extensions
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Helper to sanitize table and column names for SQL safety
def sanitize_name(name):
    # Keep alphanumeric characters and underscores, strip others
    sanitized = re.sub(r'[^a-zA-Z0-9_]', '_', name.strip())
    # Ensure it starts with a letter or underscore
    if sanitized and sanitized[0].isdigit():
        sanitized = '_' + sanitized
    return sanitized or 'table_col'

# Helper to guarantee unique column names from raw file headers
def make_unique_column_names(column_names):
    seen = {}
    unique_names = []
    for i, name in enumerate(column_names):
        name_str = str(name).strip() if name is not None else f"Column_{i}"
        sanitized = sanitize_name(name_str)
        if not sanitized:
            sanitized = f"column_{i}"
            
        key = sanitized.lower()
        if key in seen:
            seen[key] += 1
            suffix = seen[key]
            candidate = f"{sanitized}_{suffix}"
            while candidate.lower() in seen:
                suffix += 1
                candidate = f"{sanitized}_{suffix}"
            seen[candidate.lower()] = 0
            unique_names.append(candidate)
        else:
            seen[key] = 0
            unique_names.append(sanitized)
# Helper to normalize values in Python for relation analysis
def normalize_val(val):
    if val is None:
        return ""
    val_str = str(val).strip().lower()
    if not val_str:
        return ""
    # Try to parse as float to normalize numeric strings (e.g. "1.0" -> "1", "01" -> "1")
    try:
        # Avoid treating general text as float (e.g. "12a" or "1.2.3")
        if val_str.replace('.', '', 1).replace('-', '', 1).isdigit():
            val_float = float(val_str)
            if val_float.is_integer():
                return str(int(val_float))
            return str(val_float)
    except Exception:
        pass
    return val_str

# Helper to generate robust SQL expression for column comparison
def sql_clean_column(col_name):
    return f"""LOWER(TRIM(
        CASE 
            WHEN "{col_name}" NOT GLOB '*[^0-9.-]*' AND "{col_name}" GLOB '*[0-9]*' THEN
                CASE 
                    WHEN CAST("{col_name}" AS REAL) = CAST("{col_name}" AS INTEGER) THEN CAST(CAST("{col_name}" AS INTEGER) AS TEXT)
                    ELSE CAST(CAST("{col_name}" AS REAL) AS TEXT)
                END
            ELSE 
                CAST("{col_name}" AS TEXT)
        END
    ))"""

# Helper to translate database errors to user-friendly messages in Turkish
def translate_db_error(e):
    err_msg = str(e)
    
    # UNIQUE veya PRIMARY KEY kısıtlaması hatası
    if "UNIQUE constraint failed" in err_msg:
        parts = err_msg.split(":")
        col_detail = parts[-1].strip() if len(parts) > 1 else ""
        display_col = col_detail.replace("_new", "")
        return (
            f"Benzersizlik (Unique) Hatası: '{display_col}' sütununda zaten mevcut olan (yinelenen) bir değer eklenmeye veya oluşturulmaya çalışıldı. "
            f"İlişki kurmaya çalışıyorsanız: Tanımladığınız ana tablonun (parent table) ana sütunundaki verilerin tamamen benzersiz olması gerekir. "
            f"Lütfen '{display_col}' sütunundaki yinelenen verileri temizleyin ya da ilişkiyi kurarken Ana Tablo ile İlişkili Tablo seçimlerinin yönünü kontrol edin."
        )
        
    # FOREIGN KEY kısıtlaması hatası
    if "FOREIGN KEY constraint failed" in err_msg:
        return (
            "İlişki (Foreign Key) Hatası: İlişkili tablolardaki veri bütünlüğü kuralı ihlal edildi. "
            "Girdiğiniz değer ana tabloda mevcut olmayabilir ya da silmek/güncellemek istediğiniz veri diğer bir tablo tarafından referans alınıyor olabilir."
        )
        
    # NOT NULL kısıtlaması hatası
    if "NOT NULL constraint failed" in err_msg:
        parts = err_msg.split(":")
        col_detail = parts[-1].strip() if len(parts) > 1 else ""
        display_col = col_detail.replace("_new", "")
        return f"Boş Bırakılamaz Hatası: '{display_col}' sütunu boş bırakılamaz. Lütfen bu alanı doldurun."
        
    return err_msg

# Helper to guess data type from a list of values
def guess_data_type(values):
    if not values:
        return 'TEXT'
    
    is_int = True
    is_real = True
    
    non_empty_count = 0
    for val in values:
        if val is None or val == '':
            continue
        non_empty_count += 1
        
        # Check Integer
        if is_int:
            try:
                int(str(val))
            except ValueError:
                is_int = False
                
        # Check Float (Real)
        if is_real:
            try:
                float(str(val))
            except ValueError:
                is_real = False
                
    if non_empty_count == 0:
        return 'TEXT'
    if is_int:
        return 'INTEGER'
    if is_real:
        return 'REAL'
    return 'TEXT'

# --- ROUTING FOR FRONTEND ---
@app.route('/')
def index():
    return app.send_static_file('index.html')

# --- API ENDPOINTS ---

# 1. List all tables
@app.route('/api/tables', methods=['GET'])
def list_tables():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Query user-defined tables in SQLite
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%';")
        tables = [row['name'] for row in cursor.fetchall()]
        
        result = []
        for table in tables:
            # Get column info
            cursor.execute(f"PRAGMA table_info({table});")
            columns = [{'name': col['name'], 'type': col['type']} for col in cursor.fetchall()]
            
            # Get row count
            cursor.execute(f"SELECT COUNT(*) as count FROM {table};")
            row_count = cursor.fetchone()['count']
            
            result.append({
                'name': table,
                'columns': columns,
                'rowCount': row_count
            })
            
        conn.close()
        return jsonify({'success': True, 'tables': result})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# 2. Create table manually
@app.route('/api/tables', methods=['POST'])
def create_table():
    try:
        data = request.json
        table_name = sanitize_name(data.get('table_name'))
        columns_data = data.get('columns', [])
        
        if not table_name:
            return jsonify({'success': False, 'error': 'Table name is required.'}), 400
        if not columns_data:
            return jsonify({'success': False, 'error': 'At least one column is required.'}), 400
            
        column_defs = []
        for col in columns_data:
            col_name = sanitize_name(col.get('name'))
            col_type = col.get('type', 'TEXT').upper()
            if col_type not in ['TEXT', 'INTEGER', 'REAL', 'NUMERIC', 'BLOB', 'DATE', 'DATETIME', 'FLOAT', 'DOUBLE', 'VARCHAR', 'BOOLEAN']:
                col_type = 'TEXT'
            column_defs.append(f"{col_name} {col_type}")
            
        sql = f"CREATE TABLE {table_name} ({', '.join(column_defs)});"
        
        conn = get_db_connection()
        conn.execute(sql)
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': f"Table '{table_name}' created successfully."})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# 3. Delete table
@app.route('/api/tables/<table_name>', methods=['DELETE'])
def delete_table(table_name):
    try:
        table_name = sanitize_name(table_name)
        conn = get_db_connection()
        conn.execute(f"DROP TABLE {table_name};")
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': f"Table '{table_name}' deleted successfully."})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# 4. Get table data (paginated, sorted, searched)
@app.route('/api/tables/<table_name>', methods=['GET'])
def get_table_data(table_name):
    try:
        table_name = sanitize_name(table_name)
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 50))
        search = request.args.get('search', '').strip()
        sort_by = request.args.get('sort_by', '').strip()
        sort_order = request.args.get('sort_order', 'asc').lower()
        
        # Relation filters
        rel_filter_type = request.args.get('rel_filter_type', '').strip()
        rel_filter_table = request.args.get('rel_filter_table', '').strip()
        rel_filter_col = request.args.get('rel_filter_col', '').strip()
        rel_filter_other_col = request.args.get('rel_filter_other_col', '').strip()
        
        if sort_order not in ['asc', 'desc']:
            sort_order = 'asc'
            
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Get column information
        cursor.execute(f"PRAGMA table_info({table_name});")
        columns_info = cursor.fetchall()
        columns = [col['name'] for col in columns_info]
        column_types = {col['name']: col['type'] for col in columns_info}
        
        # Build search condition
        where_clauses = []
        query_params = []
        
        if search:
            search_conditions = []
            for col in columns:
                search_conditions.append(f"CAST({col} AS TEXT) LIKE ?")
                query_params.append(f"%{search}%")
            where_clauses.append("(" + " OR ".join(search_conditions) + ")")
            
        # Sütun bazlı filtreleme (Column-specific filters)
        for col in columns:
            col_filter = request.args.get(f"filter_{col}", "").strip()
            if col_filter:
                where_clauses.append(f"CAST({col} AS TEXT) LIKE ?")
                query_params.append(f"%{col_filter}%")
                
        # İlişki bazlı filtreleme (Relation filters: matched/unmatched)
        if rel_filter_type and rel_filter_table and rel_filter_col and rel_filter_other_col:
            rel_filter_table = sanitize_name(rel_filter_table)
            rel_filter_col = sanitize_name(rel_filter_col)
            rel_filter_other_col = sanitize_name(rel_filter_other_col)
            
            clean_col = sql_clean_column(rel_filter_col)
            clean_other = sql_clean_column(rel_filter_other_col)
            
            if rel_filter_type == 'matched':
                where_clauses.append(f"{clean_col} IN (SELECT {clean_other} FROM \"{rel_filter_table}\")")
            elif rel_filter_type == 'unmatched':
                where_clauses.append(f"(\"{rel_filter_col}\" IS NULL OR {clean_col} NOT IN (SELECT {clean_other} FROM \"{rel_filter_table}\" WHERE \"{rel_filter_other_col}\" IS NOT NULL))")
                
        where_clause = ""
        if where_clauses:
            where_clause = " WHERE " + " AND ".join(where_clauses)
            
        # Get total count
        count_sql = f"SELECT COUNT(*) as count FROM {table_name}{where_clause}"
        cursor.execute(count_sql, query_params)
        total_rows = cursor.fetchone()['count']
        
        # Build sort clause
        sort_clause = ""
        if sort_by and sort_by in columns:
            sort_clause = f" ORDER BY {sort_by} {sort_order.upper()}"
        else:
            # Default to ordering by rowid to preserve insertion order
            sort_clause = " ORDER BY rowid ASC"
            
        # Pagination
        offset = (page - 1) * per_page
        limit_clause = f" LIMIT ? OFFSET ?"
        
        # Get rows
        # Select rowid explicitly so we can perform CRUD operations on individual records
        select_sql = f"SELECT rowid as _rowid_, * FROM {table_name}{where_clause}{sort_clause}{limit_clause}"
        cursor.execute(select_sql, query_params + [per_page, offset])
        rows = [dict(row) for row in cursor.fetchall()]
        
        conn.close()
        
        return jsonify({
            'success': True,
            'table': table_name,
            'columns': columns,
            'column_types': column_types,
            'rows': rows,
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total_rows': total_rows,
                'total_pages': (total_rows + per_page - 1) // per_page if total_rows > 0 else 1
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# 4b. Get all rowids (matching current search and column filters)
@app.route('/api/tables/<table_name>/rowids', methods=['GET'])
def get_table_rowids(table_name):
    try:
        table_name = sanitize_name(table_name)
        search = request.args.get('search', '').strip()
        
        # Relation filters
        rel_filter_type = request.args.get('rel_filter_type', '').strip()
        rel_filter_table = request.args.get('rel_filter_table', '').strip()
        rel_filter_col = request.args.get('rel_filter_col', '').strip()
        rel_filter_other_col = request.args.get('rel_filter_other_col', '').strip()
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Get column information
        cursor.execute(f"PRAGMA table_info({table_name});")
        columns_info = cursor.fetchall()
        columns = [col['name'] for col in columns_info]
        
        # Build search condition
        where_clauses = []
        query_params = []
        
        if search:
            search_conditions = []
            for col in columns:
                search_conditions.append(f"CAST({col} AS TEXT) LIKE ?")
                query_params.append(f"%{search}%")
            where_clauses.append("(" + " OR ".join(search_conditions) + ")")
            
        # Sütun bazlı filtreleme (Column-specific filters)
        for col in columns:
            col_filter = request.args.get(f"filter_{col}", "").strip()
            if col_filter:
                where_clauses.append(f"CAST({col} AS TEXT) LIKE ?")
                query_params.append(f"%{col_filter}%")
                
        # İlişki bazlı filtreleme (Relation filters: matched/unmatched)
        if rel_filter_type and rel_filter_table and rel_filter_col and rel_filter_other_col:
            rel_filter_table = sanitize_name(rel_filter_table)
            rel_filter_col = sanitize_name(rel_filter_col)
            rel_filter_other_col = sanitize_name(rel_filter_other_col)
            
            clean_col = sql_clean_column(rel_filter_col)
            clean_other = sql_clean_column(rel_filter_other_col)
            
            if rel_filter_type == 'matched':
                where_clauses.append(f"{clean_col} IN (SELECT {clean_other} FROM \"{rel_filter_table}\")")
            elif rel_filter_type == 'unmatched':
                where_clauses.append(f"(\"{rel_filter_col}\" IS NULL OR {clean_col} NOT IN (SELECT {clean_other} FROM \"{rel_filter_table}\" WHERE \"{rel_filter_other_col}\" IS NOT NULL))")
                
        where_clause = ""
        if where_clauses:
            where_clause = " WHERE " + " AND ".join(where_clauses)
            
        # Get rowids
        sql = f"SELECT rowid as _rowid_ FROM {table_name}{where_clause}"
        cursor.execute(sql, query_params)
        rowids = [row['_rowid_'] for row in cursor.fetchall()]
        conn.close()
        
        return jsonify({
            'success': True,
            'rowids': rowids
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# 5. Insert row
@app.route('/api/tables/<table_name>/row', methods=['POST'])
def insert_row(table_name):
    try:
        table_name = sanitize_name(table_name)
        row_data = request.json
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Get actual table columns
        cursor.execute(f"PRAGMA table_info({table_name});")
        columns = [col['name'] for col in cursor.fetchall()]
        
        # Filter input data to match existing columns
        insert_cols = []
        insert_vals = []
        for col in columns:
            if col in row_data:
                insert_cols.append(col)
                insert_vals.append(row_data[col])
                
        if not insert_cols:
            return jsonify({'success': False, 'error': 'No valid columns provided for insertion.'}), 400
            
        sql = f"INSERT INTO {table_name} ({', '.join(insert_cols)}) VALUES ({', '.join(['?'] * len(insert_vals))})"
        cursor.execute(sql, insert_vals)
        conn.commit()
        
        new_rowid = cursor.lastrowid
        conn.close()
        
        return jsonify({'success': True, 'message': 'Row inserted successfully.', 'rowid': new_rowid})
    except Exception as e:
        return jsonify({'success': False, 'error': translate_db_error(e)}), 500

# 6. Update row
@app.route('/api/tables/<table_name>/row/<int:rowid>', methods=['PUT'])
def update_row(table_name, rowid):
    try:
        table_name = sanitize_name(table_name)
        row_data = request.json
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Get actual table columns
        cursor.execute(f"PRAGMA table_info({table_name});")
        columns = [col['name'] for col in cursor.fetchall()]
        
        update_clauses = []
        update_vals = []
        for col in columns:
            if col in row_data:
                update_clauses.append(f"{col} = ?")
                update_vals.append(row_data[col])
                
        if not update_clauses:
            return jsonify({'success': False, 'error': 'No columns to update.'}), 400
            
        sql = f"UPDATE {table_name} SET {', '.join(update_clauses)} WHERE rowid = ?"
        cursor.execute(sql, update_vals + [rowid])
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Row updated successfully.'})
    except Exception as e:
        return jsonify({'success': False, 'error': translate_db_error(e)}), 500

# 7. Delete row
@app.route('/api/tables/<table_name>/row/<int:rowid>', methods=['DELETE'])
def delete_row(table_name, rowid):
    try:
        table_name = sanitize_name(table_name)
        conn = get_db_connection()
        conn.execute(f"DELETE FROM {table_name} WHERE rowid = ?;", (rowid,))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': 'Row deleted successfully.'})
    except Exception as e:
        return jsonify({'success': False, 'error': translate_db_error(e)}), 500

# 7a. Bulk delete rows
@app.route('/api/tables/<table_name>/rows/bulk-delete', methods=['POST'])
def bulk_delete_rows(table_name):
    try:
        table_name = sanitize_name(table_name)
        data = request.json
        rowids = data.get('rowids', [])
        
        if not rowids:
            return jsonify({'success': False, 'error': 'No rows selected for deletion.'}), 400
            
        placeholders = ', '.join(['?'] * len(rowids))
        sql = f"DELETE FROM {table_name} WHERE rowid IN ({placeholders});"
        
        conn = get_db_connection()
        conn.execute(sql, rowids)
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': f'{len(rowids)} rows deleted successfully.'})
    except Exception as e:
        return jsonify({'success': False, 'error': translate_db_error(e)}), 500

# 7b. Bulk update rows
@app.route('/api/tables/<table_name>/rows/bulk-update', methods=['POST'])
def bulk_update_rows(table_name):
    try:
        table_name = sanitize_name(table_name)
        data = request.json
        rowids = data.get('rowids', [])
        column = sanitize_name(data.get('column', ''))
        value = data.get('value', '')
        
        if not rowids:
            return jsonify({'success': False, 'error': 'No rows selected for update.'}), 400
        if not column:
            return jsonify({'success': False, 'error': 'Column name is required.'}), 400
            
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(f"PRAGMA table_info({table_name});")
        columns = [col['name'] for col in cursor.fetchall()]
        
        if column not in columns:
            conn.close()
            return jsonify({'success': False, 'error': f"Column '{column}' does not exist in table."}), 400
            
        placeholders = ', '.join(['?'] * len(rowids))
        sql = f"UPDATE {table_name} SET {column} = ? WHERE rowid IN ({placeholders});"
        
        cursor.execute(sql, [value] + rowids)
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': f'{len(rowids)} rows updated successfully.'})
    except Exception as e:
        return jsonify({'success': False, 'error': translate_db_error(e)}), 500

# 8. Parse uploaded CSV/Excel file for preview
@app.route('/api/parse-file', methods=['POST'])
def parse_file():
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file part in the request.'}), 400
            
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected.'}), 400
            
        if not allowed_file(file.filename):
            return jsonify({'success': False, 'error': 'Invalid file format. Allowed formats: CSV, XLSX, XLS.'}), 400
            
        # Save file to temp directory
        file_ext = file.filename.rsplit('.', 1)[1].lower()
        file_key = str(uuid.uuid4())
        temp_filename = f"{file_key}.{file_ext}"
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], temp_filename)
        file.save(file_path)
        
        suggested_table_name = sanitize_name(file.filename.rsplit('.', 1)[0])
        
        # Read metadata based on extension
        sheets = []
        columns = []
        preview_rows = []
        
        if file_ext in ['xlsx', 'xls']:
            try:
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                sheets = wb.sheetnames
                
                # Preview first sheet by default
                if sheets:
                    sheet = wb[sheets[0]]
                    row_generator = sheet.iter_rows(values_only=True)
                    
                    # Find header row (first non-empty row)
                    header = None
                    rows_data = []
                    for row in row_generator:
                        # Check if row has any non-none value
                        if any(cell is not None for cell in row):
                            if not header:
                                header = [str(cell) if cell is not None else f"Column_{i}" for i, cell in enumerate(row)]
                            else:
                                rows_data.append(row)
                                if len(rows_data) >= 5:
                                    break
                    
                    if header:
                        columns = make_unique_column_names(header)
                        
                        # Transpose for data type guessing
                        col_values = {col: [] for col in columns}
                        for r in rows_data:
                            for idx, val in enumerate(r):
                                if idx < len(columns):
                                    col_values[columns[idx]].append(val)
                                    
                        detected_types = {col: guess_data_type(col_values[col]) for col in columns}
                        
                        # Build column structures
                        columns_meta = [{'name': col, 'type': detected_types[col]} for col in columns]
                        
                        # Prepare row dictionaries for preview
                        for r in rows_data:
                            row_dict = {}
                            for idx, val in enumerate(r):
                                if idx < len(columns):
                                    # Convert non-serializable objects (like datetime) to string
                                    row_dict[columns[idx]] = str(val) if val is not None else ''
                            preview_rows.append(row_dict)
                            
                        columns = columns_meta
                wb.close()
            except Exception as e:
                # Cleanup if parse failed
                if os.path.exists(file_path):
                    os.remove(file_path)
                return jsonify({'success': False, 'error': f"Failed to parse Excel file: {str(e)}"}), 500
        else:
            # CSV file
            try:
                with open(file_path, 'r', encoding='utf-8-sig', errors='ignore') as f:
                    reader = csv.reader(f)
                    header = next(reader, None)
                    if header:
                        columns = make_unique_column_names(header)
                        
                        rows_data = []
                        for i, r in enumerate(reader):
                            rows_data.append(r)
                            if len(rows_data) >= 5:
                                break
                                
                        col_values = {col: [] for col in columns}
                        for r in rows_data:
                            for idx, val in enumerate(r):
                                if idx < len(columns):
                                    col_values[columns[idx]].append(val)
                                    
                        detected_types = {col: guess_data_type(col_values[col]) for col in columns}
                        columns_meta = [{'name': col, 'type': detected_types[col]} for col in columns]
                        
                        for r in rows_data:
                            row_dict = {}
                            for idx, val in enumerate(r):
                                if idx < len(columns):
                                    row_dict[columns[idx]] = val
                            preview_rows.append(row_dict)
                            
                        columns = columns_meta
            except Exception as e:
                # Cleanup if parse failed
                if os.path.exists(file_path):
                    os.remove(file_path)
                return jsonify({'success': False, 'error': f"Failed to parse CSV file: {str(e)}"}), 500
                
        return jsonify({
            'success': True,
            'file_key': file_key,
            'file_ext': file_ext,
            'suggested_table_name': suggested_table_name,
            'sheets': sheets,
            'columns': columns,
            'preview_rows': preview_rows
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# 8b. Preview Excel sheet change
@app.route('/api/preview-sheet', methods=['GET'])
def preview_sheet():
    try:
        file_key = request.args.get('file_key')
        file_ext = request.args.get('file_ext')
        sheet_name = request.args.get('sheet_name')
        
        if not file_key or not file_ext or not sheet_name:
            return jsonify({'success': False, 'error': 'Missing parameters.'}), 400
            
        temp_filename = f"{file_key}.{file_ext}"
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], temp_filename)
        
        if not os.path.exists(file_path):
            return jsonify({'success': False, 'error': 'Uploaded file not found.'}), 400
            
        columns = []
        preview_rows = []
        
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            return jsonify({'success': False, 'error': f"Sheet '{sheet_name}' not found."}), 400
            
        sheet = wb[sheet_name]
        row_generator = sheet.iter_rows(values_only=True)
        
        header = None
        rows_data = []
        for row in row_generator:
            if any(cell is not None for cell in row):
                if not header:
                    header = [str(cell) if cell is not None else f"Column_{i}" for i, cell in enumerate(row)]
                else:
                    rows_data.append(row)
                    if len(rows_data) >= 5:
                        break
        
        if header:
            columns = make_unique_column_names(header)
            
            col_values = {col: [] for col in columns}
            for r in rows_data:
                for idx, val in enumerate(r):
                    if idx < len(columns):
                        col_values[columns[idx]].append(val)
                        
            detected_types = {col: guess_data_type(col_values[col]) for col in columns}
            columns_meta = [{'name': col, 'type': detected_types[col]} for col in columns]
            
            for r in rows_data:
                row_dict = {}
                for idx, val in enumerate(r):
                    if idx < len(columns):
                        row_dict[columns[idx]] = str(val) if val is not None else ''
                preview_rows.append(row_dict)
                
            columns = columns_meta
        wb.close()
            
        return jsonify({
            'success': True,
            'columns': columns,
            'preview_rows': preview_rows
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# 9. Import file contents to SQLite
@app.route('/api/import-file', methods=['POST'])
def import_file():
    file_path = None
    try:
        data = request.json
        file_key = data.get('file_key')
        file_ext = data.get('file_ext')
        sheet_name = data.get('sheet_name')
        table_name = sanitize_name(data.get('table_name'))
        import_mode = data.get('import_mode', 'new')  # 'new' or 'append'
        column_mapping = data.get('column_mapping', {}) # file_col -> db_col (for append)
        columns_def = data.get('columns', []) # list of {name, type} for table creation
        
        if not file_key or not file_ext:
            return jsonify({'success': False, 'error': 'Invalid file upload parameters.'}), 400
            
        temp_filename = f"{file_key}.{file_ext}"
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], temp_filename)
        
        if not os.path.exists(file_path):
            return jsonify({'success': False, 'error': 'Uploaded file has expired or was not found.'}), 400
            
        # Parse all rows from file
        rows_to_insert = []
        file_headers = []
        
        if file_ext in ['xlsx', 'xls']:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            if not sheet_name:
                sheet_name = wb.sheetnames[0]
            sheet = wb[sheet_name]
            
            row_generator = sheet.iter_rows(values_only=True)
            for row in row_generator:
                if any(cell is not None for cell in row):
                    if not file_headers:
                        file_headers = make_unique_column_names(row)
                    else:
                        row_dict = {}
                        for idx, val in enumerate(row):
                            if idx < len(file_headers):
                                # Convert dates/times to string for DB storage
                                row_dict[file_headers[idx]] = val
                        rows_to_insert.append(row_dict)
            wb.close()
        else:
            # CSV
            with open(file_path, 'r', encoding='utf-8-sig', errors='ignore') as f:
                reader = csv.reader(f)
                header = next(reader, None)
                if header:
                    file_headers = make_unique_column_names(header)
                    for r in reader:
                        row_dict = {}
                        for idx, val in enumerate(r):
                            if idx < len(file_headers):
                                row_dict[file_headers[idx]] = val
                        rows_to_insert.append(row_dict)
                        
        if not file_headers:
            return jsonify({'success': False, 'error': 'No header row found in file.'}), 400
            
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Check if table already exists
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name = ?;", (table_name,))
        table_exists = cursor.fetchone() is not None
        
        if import_mode == 'new':
            if table_exists:
                conn.close()
                return jsonify({'success': False, 'error': f"Table '{table_name}' already exists. Choose a different name or use append mode."}), 400
                
            # Create Table structure
            col_defs = []
            for col in columns_def:
                c_name = sanitize_name(col.get('name'))
                c_type = col.get('type', 'TEXT').upper()
                col_defs.append(f"{c_name} {c_type}")
                
            create_sql = f"CREATE TABLE {table_name} ({', '.join(col_defs)});"
            cursor.execute(create_sql)
            
            # For new table, columns to insert are simple
            db_columns = [sanitize_name(col.get('name')) for col in columns_def]
            
            # Insert values
            insert_sql = f"INSERT INTO {table_name} ({', '.join(db_columns)}) VALUES ({', '.join(['?'] * len(db_columns))})"
            
            for r in rows_to_insert:
                val_list = []
                for db_col in db_columns:
                    # Find corresponding header in file
                    val = r.get(db_col, None)
                    val_list.append(val)
                cursor.execute(insert_sql, val_list)
                
        else: # Append mode
            if not table_exists:
                conn.close()
                return jsonify({'success': False, 'error': f"Table '{table_name}' does not exist."}), 400
                
            # Get existing table columns
            cursor.execute(f"PRAGMA table_info({table_name});")
            existing_db_cols = [col['name'] for col in cursor.fetchall()]
            
            # Validate mapping: user maps file header -> database column
            # column_mapping is a dict { file_header: db_column }
            insert_cols = []
            for file_col, db_col in column_mapping.items():
                if db_col in existing_db_cols and file_col in file_headers:
                    insert_cols.append((file_col, db_col))
                    
            if not insert_cols:
                conn.close()
                return jsonify({'success': False, 'error': 'No matching columns found or mapped.'}), 400
                
            db_cols_to_insert = [mapped[1] for mapped in insert_cols]
            file_cols_to_insert = [mapped[0] for mapped in insert_cols]
            
            insert_sql = f"INSERT INTO {table_name} ({', '.join(db_cols_to_insert)}) VALUES ({', '.join(['?'] * len(db_cols_to_insert))})"
            
            for r in rows_to_insert:
                val_list = []
                for file_col in file_cols_to_insert:
                    val_list.append(r.get(file_col, None))
                cursor.execute(insert_sql, val_list)
                
        conn.commit()
        conn.close()
        
        # Cleanup temp file
        if os.path.exists(file_path):
            os.remove(file_path)
            
        return jsonify({'success': True, 'message': f"Imported {len(rows_to_insert)} rows into table '{table_name}' successfully."})
        
    except Exception as e:
        # Cleanup temp file on failure
        if file_path and os.path.exists(file_path):
            os.remove(file_path)
        return jsonify({'success': False, 'error': translate_db_error(e)}), 500

# 10. Export database table as CSV or Excel
@app.route('/api/tables/<table_name>/export', methods=['GET'])
def export_table(table_name):
    try:
        table_name = sanitize_name(table_name)
        export_format = request.args.get('format', 'csv').lower()
        search = request.args.get('search', '').strip()
        sort_by = request.args.get('sort_by', '').strip()
        sort_order = request.args.get('sort_order', 'asc').lower()
        
        # Relation filters
        rel_filter_type = request.args.get('rel_filter_type', '').strip()
        rel_filter_table = request.args.get('rel_filter_table', '').strip()
        rel_filter_col = request.args.get('rel_filter_col', '').strip()
        rel_filter_other_col = request.args.get('rel_filter_other_col', '').strip()
        
        if sort_order not in ['asc', 'desc']:
            sort_order = 'asc'
            
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Get column names
        cursor.execute(f"PRAGMA table_info({table_name});")
        columns_info = cursor.fetchall()
        columns = [col['name'] for col in columns_info]
        
        if not columns:
            conn.close()
            return "Table not found", 404
            
        # Build search condition
        where_clauses = []
        query_params = []
        
        if search:
            search_conditions = []
            for col in columns:
                search_conditions.append(f"CAST({col} AS TEXT) LIKE ?")
                query_params.append(f"%{search}%")
            where_clauses.append("(" + " OR ".join(search_conditions) + ")")
            
        # Sütun bazlı filtreleme (Column-specific filters)
        for col in columns:
            col_filter = request.args.get(f"filter_{col}", "").strip()
            if col_filter:
                where_clauses.append(f"CAST({col} AS TEXT) LIKE ?")
                query_params.append(f"%{col_filter}%")
                
        # İlişki bazlı filtreleme (Relation filters: matched/unmatched)
        if rel_filter_type and rel_filter_table and rel_filter_col and rel_filter_other_col:
            rel_filter_table = sanitize_name(rel_filter_table)
            rel_filter_col = sanitize_name(rel_filter_col)
            rel_filter_other_col = sanitize_name(rel_filter_other_col)
            
            clean_col = sql_clean_column(rel_filter_col)
            clean_other = sql_clean_column(rel_filter_other_col)
            
            if rel_filter_type == 'matched':
                where_clauses.append(f"{clean_col} IN (SELECT {clean_other} FROM \"{rel_filter_table}\")")
            elif rel_filter_type == 'unmatched':
                where_clauses.append(f"(\"{rel_filter_col}\" IS NULL OR {clean_col} NOT IN (SELECT {clean_other} FROM \"{rel_filter_table}\" WHERE \"{rel_filter_other_col}\" IS NOT NULL))")
                
        where_clause = ""
        if where_clauses:
            where_clause = " WHERE " + " AND ".join(where_clauses)
            
        # Build sort clause
        sort_clause = ""
        if sort_by and sort_by in columns:
            sort_clause = f" ORDER BY {sort_by} {sort_order.upper()}"
        else:
            # Default to ordering by rowid to preserve insertion order
            sort_clause = " ORDER BY rowid ASC"
            
        cursor.execute(f"SELECT * FROM {table_name}{where_clause}{sort_clause};", query_params)
        rows = [list(row) for row in cursor.fetchall()]
        conn.close()
        
        export_filename = f"{table_name}_export.{export_format}"
        export_path = os.path.join(UPLOAD_FOLDER, export_filename)
        
        if export_format == 'csv':
            # Write to CSV
            with open(export_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow(columns)
                writer.writerows(rows)
                
            return send_file(
                export_path,
                mimetype='text/csv',
                as_attachment=True,
                download_name=export_filename
            )
            
        elif export_format == 'excel':
            # Write to Excel using openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = table_name[:30] # Excel limit sheet title length
            
            # Write header
            ws.append(columns)
            
            # Write rows
            for r in rows:
                ws.append(r)
                
            wb.save(export_path)
            
            return send_file(
                export_path,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=export_filename
            )
        else:
            return "Invalid export format. Use 'csv' or 'excel'.", 400
            
    except Exception as e:
        return f"Export failed: {str(e)}", 500

def rebuild_table_schema(table_name, col_modifiers=None, new_fks=None, remove_fk=None):
    if col_modifiers is None:
        col_modifiers = {}
    conn = get_db_connection()
    # Turn off foreign keys temporarily to allow table re-creation
    conn.execute("PRAGMA foreign_keys = OFF;")
    cursor = conn.cursor()
    
    try:
        # 1. Get existing columns and primary keys
        cursor.execute(f"PRAGMA table_info({table_name});")
        cols = [dict(row) for row in cursor.fetchall()]
        if not cols:
            raise Exception(f"Table '{table_name}' does not exist.")
            
        # 2. Get existing foreign keys
        cursor.execute(f"PRAGMA foreign_key_list({table_name});")
        fks = [dict(row) for row in cursor.fetchall()]
        
        # Group existing foreign keys by relation ID
        fk_groups = {}
        for fk in fks:
            fk_id = fk['id']
            if fk_id not in fk_groups:
                fk_groups[fk_id] = {
                    'table': fk['table'],
                    'from': [],
                    'to': [],
                    'on_update': fk['on_update'],
                    'on_delete': fk['on_delete']
                }
            fk_groups[fk_id]['from'].append(fk['from'])
            fk_groups[fk_id]['to'].append(fk['to'])
            
        # 3. Reconstruct column definitions
        # Check if parent_column is being modified
        has_existing_pk = any(col['pk'] > 0 for col in cols)
        
        col_defs = []
        inline_pk = False
        
        # We check if there's a single integer PK to declare it inline
        pk_cols = [c for c in cols if c['pk'] > 0]
        if len(pk_cols) == 1 and pk_cols[0]['type'].upper() == 'INTEGER':
            inline_pk = True
            
        for col in cols:
            c_name = col['name']
            c_type = col['type']
            
            # Check if there is a modifier for this column
            modifier = col_modifiers.get(c_name)
            
            if modifier:
                col_def = f"{c_name} {c_type} {modifier}"
            else:
                col_def = f"{c_name} {c_type}"
                if col['pk'] > 0 and inline_pk:
                    col_def += " PRIMARY KEY"
                if col['notnull']:
                    col_def += " NOT NULL"
                if col['dflt_value'] is not None:
                    col_def += f" DEFAULT {col['dflt_value']}"
            col_defs.append(col_def)
            
        if len(pk_cols) > 0 and not inline_pk:
            pk_names = ", ".join(col['name'] for col in sorted(pk_cols, key=lambda x: x['pk']))
            col_defs.append(f"PRIMARY KEY ({pk_names})")
            
        # 4. Reconstruct foreign key clauses
        fk_clauses = []
        for fk_id, group in fk_groups.items():
            # Check if we should remove this relation
            if remove_fk:
                # remove_fk is a dict: {'child_column': x, 'parent_table': y, 'parent_column': z}
                if len(group['from']) == 1 and len(group['to']) == 1:
                    if (group['from'][0] == remove_fk['child_column'] and 
                        group['table'] == remove_fk['parent_table'] and 
                        group['to'][0] == remove_fk['parent_column']):
                        continue
            
            from_cols = ", ".join(group['from'])
            to_cols = ", ".join(group['to'])
            clause = f"FOREIGN KEY ({from_cols}) REFERENCES {group['table']} ({to_cols}) ON UPDATE {group['on_update']} ON DELETE {group['on_delete']}"
            fk_clauses.append(clause)
            
        # Add new foreign keys if any
        if new_fks:
            for new_fk in new_fks:
                on_update = new_fk.get('on_update', 'CASCADE')
                on_delete = new_fk.get('on_delete', 'CASCADE')
                clause = f"FOREIGN KEY ({new_fk['child_column']}) REFERENCES {new_fk['parent_table']} ({new_fk['parent_column']}) ON UPDATE {on_update} ON DELETE {on_delete}"
                fk_clauses.append(clause)
                
        # Combine column definitions and FK clauses
        table_body = col_defs + fk_clauses
        new_schema_sql = f"CREATE TABLE {table_name}_new ({', '.join(table_body)});"
        
        # 5. Fetch existing custom indices (excluding AUTO indices)
        cursor.execute("SELECT name, sql FROM sqlite_master WHERE type='index' AND tbl_name = ? AND sql IS NOT NULL;", (table_name,))
        indices = [dict(row) for row in cursor.fetchall()]
        
        # 6. Execute modification transaction
        cursor.execute("BEGIN TRANSACTION;")
        
        # Create new table
        cursor.execute(new_schema_sql)
        
        # Copy data
        columns_str = ", ".join(col['name'] for col in cols)
        cursor.execute(f"INSERT INTO {table_name}_new ({columns_str}) SELECT {columns_str} FROM {table_name};")
        
        # Drop old table and rename new table
        cursor.execute(f"DROP TABLE {table_name};")
        cursor.execute(f"ALTER TABLE {table_name}_new RENAME TO {table_name};")
        
        # Recreate indices
        for idx in indices:
            cursor.execute(idx['sql'])
            
        # Verify foreign keys before committing!
        cursor.execute("PRAGMA foreign_key_check;")
        violations = cursor.fetchall()
        if violations:
            viol_details = []
            for v in violations[:5]:
                child_table = v[0]
                row_id = v[1]
                parent_table = v[2]
                fk_id = v[3]
                
                # Query foreign key details to identify the child column name
                cursor.execute(f"PRAGMA foreign_key_list({child_table});")
                fk_list = [dict(row) for row in cursor.fetchall()]
                child_col = None
                for fk in fk_list:
                    if fk['id'] == fk_id:
                        child_col = fk['from']
                        break
                
                if child_col:
                    cursor.execute(f"SELECT {child_col} FROM {child_table} WHERE rowid = ?;", (row_id,))
                    val_row = cursor.fetchone()
                    viol_val = val_row[0] if val_row else 'Bilinmeyen'
                    viol_details.append(
                        f"'{child_table}.{child_col}' sütunundaki '{viol_val}' değeri, "
                        f"Ana '{parent_table}' tablosunda bulunmamaktadır."
                    )
                else:
                    viol_details.append(f"Tablo '{child_table}' satır ID {row_id} -> '{parent_table}' tablosunda referans bulunamadı.")
            
            if len(violations) > 5:
                viol_details.append(f"...ve {len(violations) - 5} adet daha uyuşmayan kayıt var.")
                
            raise Exception(
                "Yabancı anahtar kısıtlaması ihlal edildi. "
                "İlişkili (Child) tablodaki bazı veriler Ana (Parent) tabloda mevcut değildir:\n" + 
                "\n".join(viol_details)
            )
            
        conn.commit()
        return True
        
    except Exception as e:
        conn.rollback()
        raise e
    finally:
        conn.execute("PRAGMA foreign_keys = ON;")
        conn.close()

# 11. List all relations
@app.route('/api/relations', methods=['GET'])
def get_relations():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%';")
        tables = [row['name'] for row in cursor.fetchall()]
        
        relations = []
        for table in tables:
            cursor.execute(f"PRAGMA foreign_key_list({table});")
            fks = cursor.fetchall()
            for fk in fks:
                relations.append({
                    'child_table': table,
                    'child_column': fk['from'],
                    'parent_table': fk['table'],
                    'parent_column': fk['to'],
                    'on_update': fk['on_update'],
                    'on_delete': fk['on_delete'],
                    'is_logical': False
                })
                
        # Load virtual/logical relations
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='_sys_logical_relations';")
        if cursor.fetchone():
            cursor.execute("SELECT parent_table, parent_column, child_table, child_column, on_update, on_delete FROM _sys_logical_relations;")
            logical_fks = cursor.fetchall()
            for row in logical_fks:
                # Only include if both tables still exist in the database
                if row['parent_table'] in tables and row['child_table'] in tables:
                    relations.append({
                        'child_table': row['child_table'],
                        'child_column': row['child_column'],
                        'parent_table': row['parent_table'],
                        'parent_column': row['parent_column'],
                        'on_update': row['on_update'],
                        'on_delete': row['on_delete'],
                        'is_logical': True
                    })
                    
        conn.close()
        return jsonify({'success': True, 'relations': relations})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# 12. Create a relation
@app.route('/api/relations', methods=['POST'])
def create_relation():
    try:
        data = request.json
        parent_table = sanitize_name(data.get('parent_table'))
        parent_column = sanitize_name(data.get('parent_column'))
        child_table = sanitize_name(data.get('child_table'))
        child_column = sanitize_name(data.get('child_column'))
        on_update = data.get('on_update', 'CASCADE').upper()
        on_delete = data.get('on_delete', 'CASCADE').upper()
        is_logical = data.get('is_logical', False)
        
        if on_update not in ['CASCADE', 'SET NULL', 'RESTRICT', 'NO ACTION', 'SET DEFAULT']:
            on_update = 'CASCADE'
        if on_delete not in ['CASCADE', 'SET NULL', 'RESTRICT', 'NO ACTION', 'SET DEFAULT']:
            on_delete = 'CASCADE'
            
        if not all([parent_table, parent_column, child_table, child_column]):
            return jsonify({'success': False, 'error': 'Tüm parametreler gereklidir.'}), 400
            
        if is_logical:
            conn = get_db_connection()
            cursor = conn.cursor()
            try:
                cursor.execute("""
                    INSERT OR REPLACE INTO _sys_logical_relations 
                    (parent_table, parent_column, child_table, child_column, on_update, on_delete)
                    VALUES (?, ?, ?, ?, ?, ?);
                """, (parent_table, parent_column, child_table, child_column, on_update, on_delete))
                conn.commit()
            except Exception as e:
                conn.close()
                return jsonify({'success': False, 'error': f"Sanal ilişki kaydedilirken hata oluştu: {str(e)}"}), 500
            conn.close()
            return jsonify({'success': True, 'message': f"'{child_table}.{child_column}' -> '{parent_table}.{parent_column}' sanal ilişkisi başarıyla kuruldu."})
            
        # Step 1: Ensure parent_column is UNIQUE or PRIMARY KEY in parent_table
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(f"PRAGMA table_info({parent_table});")
        parent_cols = [dict(row) for row in cursor.fetchall()]
        conn.close()
        
        target_col = next((c for c in parent_cols if c['name'] == parent_column), None)
        if not target_col:
            return jsonify({'success': False, 'error': f"Ana sütun '{parent_column}' '{parent_table}' tablosunda bulunamadı."}), 400
            
        # Check if it has a unique or primary key constraint
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(f"PRAGMA index_list({parent_table});")
        indices = cursor.fetchall()
        
        has_unique_constraint = target_col['pk'] > 0
        if not has_unique_constraint:
            for idx in indices:
                if idx['unique']:
                    cursor.execute(f"PRAGMA index_info({idx['name']});")
                    idx_cols = [row['name'] for row in cursor.fetchall()]
                    if len(idx_cols) == 1 and idx_cols[0] == parent_column:
                        has_unique_constraint = True
                        break
        conn.close()
        
        # If it doesn't have unique constraint, make it PRIMARY KEY or UNIQUE
        if not has_unique_constraint:
            # Check for duplicate values first to provide a friendly error
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute(f"SELECT {parent_column}, COUNT(*) as cnt FROM {parent_table} WHERE {parent_column} IS NOT NULL GROUP BY {parent_column} HAVING cnt > 1 LIMIT 5;")
            duplicates = cursor.fetchall()
            conn.close()
            
            if duplicates:
                dup_vals = ", ".join([f"'{str(row[0])}'" for row in duplicates])
                return jsonify({
                    'success': False, 
                    'error': f"Ana tablo olarak seçtiğiniz '{parent_table}' tablosunun '{parent_column}' sütununda yinelenen (kopya) veriler bulunuyor (Örn: {dup_vals}). "
                             f"Veritabanı ilişkisi kurabilmek için bu sütundaki tüm değerlerin benzersiz olması gerekir. "
                             f"Lütfen '{parent_column}' sütunundaki yinelenen verileri temizleyin veya ilişki yönünün (Ana Tablo / İlişkili Tablo) doğru olduğunu kontrol edin."
                }), 400
                
            col_modifiers = {parent_column: "PRIMARY KEY" if not any(c['pk'] > 0 for c in parent_cols) else "UNIQUE"}
            rebuild_table_schema(parent_table, col_modifiers=col_modifiers)
            
        # Step 2: Add FOREIGN KEY constraint to child_table
        new_fks = [{
            'child_column': child_column,
            'parent_table': parent_table,
            'parent_column': parent_column,
            'on_update': on_update,
            'on_delete': on_delete
        }]
        rebuild_table_schema(child_table, new_fks=new_fks)
        
        return jsonify({'success': True, 'message': f"'{child_table}.{child_column}' -> '{parent_table}.{parent_column}' ilişkisi başarıyla kuruldu."})
        
    except Exception as e:
        return jsonify({'success': False, 'error': translate_db_error(e)}), 500

# 13. Delete a relation
@app.route('/api/relations', methods=['DELETE'])
def delete_relation_route():
    try:
        data = request.json
        parent_table = sanitize_name(data.get('parent_table'))
        parent_column = sanitize_name(data.get('parent_column'))
        child_table = sanitize_name(data.get('child_table'))
        child_column = sanitize_name(data.get('child_column'))
        
        if not all([parent_table, parent_column, child_table, child_column]):
            return jsonify({'success': False, 'error': 'Tüm parametreler gereklidir.'}), 400
            
        # Check if it exists in _sys_logical_relations and delete
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='_sys_logical_relations';")
        has_logical_table = cursor.fetchone() is not None
        deleted_logical = False
        if has_logical_table:
            cursor.execute("""
                DELETE FROM _sys_logical_relations 
                WHERE parent_table = ? AND parent_column = ? AND child_table = ? AND child_column = ?;
            """, (parent_table, parent_column, child_table, child_column))
            if cursor.rowcount > 0:
                deleted_logical = True
            conn.commit()
        conn.close()
        
        if deleted_logical:
            return jsonify({'success': True, 'message': 'Sanal ilişki başarıyla kaldırıldı.'})
            
        remove_fk = {
            'child_column': child_column,
            'parent_table': parent_table,
            'parent_column': parent_column
        }
        rebuild_table_schema(child_table, remove_fk=remove_fk)
        
        return jsonify({'success': True, 'message': 'İlişki başarıyla kaldırıldı.'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# 14. Analyze and suggest matching columns based on names and data values
@app.route('/api/relations/analyze', methods=['POST'])
def analyze_relations():
    try:
        data = request.json
        table_a = sanitize_name(data.get('table_a'))
        table_b = sanitize_name(data.get('table_b'))
        
        if not table_a or not table_b:
            return jsonify({'success': False, 'error': 'İki tablo ismi gereklidir.'}), 400
            
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Get table A column info
        cursor.execute(f"PRAGMA table_info({table_a});")
        cols_a_info = [dict(row) for row in cursor.fetchall()]
        cols_a = [c['name'] for c in cols_a_info]
        
        # Get table B column info
        cursor.execute(f"PRAGMA table_info({table_b});")
        cols_b_info = [dict(row) for row in cursor.fetchall()]
        cols_b = [c['name'] for c in cols_b_info]
        
        suggestions = []
        
        if not cols_a or not cols_b:
            conn.close()
            return jsonify({'success': True, 'suggestions': []})
            
        # Fetch up to 500 rows of data from table_a and table_b to compare values
        cursor.execute(f"SELECT * FROM {table_a} LIMIT 500;")
        rows_a = [dict(row) for row in cursor.fetchall()]
        
        cursor.execute(f"SELECT * FROM {table_b} LIMIT 500;")
        rows_b = [dict(row) for row in cursor.fetchall()]
        
        conn.close()
        
        # Extract unique non-empty values per column
        vals_a = {}
        for col in cols_a:
            vals_a[col] = set(normalize_val(r[col]) for r in rows_a if r.get(col) is not None)
            vals_a[col].discard("")
            
        vals_b = {}
        for col in cols_b:
            vals_b[col] = set(normalize_val(r[col]) for r in rows_b if r.get(col) is not None)
            vals_b[col].discard("")
            
        # Compare every column from Table A with Table B
        for col_a in cols_a:
            set_a = vals_a[col_a]
            if not set_a:
                continue
                
            for col_b in cols_b:
                set_b = vals_b[col_b]
                if not set_b:
                    continue
                    
                # We want to check matching criteria:
                # 1. Names are similar or identical (e.g. col_a == col_b or one ends with other, case insensitive)
                # 2. Overlap of values in set_b that are also in set_a (foreign key matching direction)
                name_match = (col_a.lower() == col_b.lower() or 
                              col_b.lower().endswith(col_a.lower()) or 
                              col_a.lower().endswith(col_b.lower()))
                              
                # Calculate what % of values in child (B) column exist in parent (A) column
                intersection = set_b.intersection(set_a)
                overlap_percent = (len(intersection) / len(set_b)) * 100 if set_b else 0
                
                # If there's high overlap (e.g., > 50%) OR name matches exactly with some overlap
                if (name_match and overlap_percent > 0) or overlap_percent >= 80:
                    suggestions.append({
                        'parent_column': col_a,
                        'child_column': col_b,
                        'overlap_percent': round(overlap_percent, 1),
                        'matching_values_count': len(intersection),
                        'total_child_values_count': len(set_b),
                        'confidence': 'high' if (name_match and overlap_percent == 100) else 'medium'
                    })
                    
        # Sort suggestions by overlap percent and confidence
        suggestions.sort(key=lambda x: (x['overlap_percent'], x['confidence'] == 'high'), reverse=True)
        
        return jsonify({'success': True, 'suggestions': suggestions})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

if __name__ == '__main__':
    print("Excel and Database Localhost App starting...")
    app.run(debug=True, port=5000)
