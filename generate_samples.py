import csv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

def generate_csv():
    customers = [
        ['Customer_ID', 'Full_Name', 'Email', 'Total_Spent', 'Is_Active'],
        ['1001', 'Ahmet Yilmaz', 'ahmet@example.com', '1250.50', 'True'],
        ['1002', 'Ayse Demir', 'ayse@example.com', '450.00', 'True'],
        ['1003', 'Mehmet Kaya', 'mehmet@example.com', '0.00', 'False'],
        ['1004', 'Fatma Celik', 'fatma@example.com', '3200.75', 'True'],
        ['1005', 'Ali Can', 'ali@example.com', '78.20', 'True']
    ]
    with open('sample_customers.csv', 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerows(customers)
    print("sample_customers.csv generated successfully.")

def generate_xlsx():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"
    
    headers = ['Employee_ID', 'First_Name', 'Last_Name', 'Department', 'Salary', 'Hire_Date']
    data = [
        [101, 'Can', 'Ozkan', 'Engineering', 85000, '2023-01-15'],
        [102, 'Ebru', 'Sahin', 'Marketing', 62000, '2022-06-01'],
        [103, 'Gokhan', 'Tepe', 'Sales', 58000, '2024-03-10'],
        [104, 'Deniz', 'Yurt', 'HR', 55000, '2021-11-20'],
        [105, 'Selin', 'Aksoy', 'Engineering', 92000, '2020-04-05']
    ]
    
    ws.append(headers)
    for row in data:
        ws.append(row)
        
    # Styling headers for a premium look in Excel
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid') # Indigo header
    header_align = Alignment(horizontal='center', vertical='center')
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        
    wb.save('sample_employees.xlsx')
    print("sample_employees.xlsx generated successfully.")

if __name__ == '__main__':
    generate_csv()
    generate_xlsx()
